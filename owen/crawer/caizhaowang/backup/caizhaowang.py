from bs4 import BeautifulSoup
from lxml import html
import xml
import requests
from openpyxl import Workbook
import openpyxl
import os, sys
sys.path.insert(0, '../')
import config
# 定义爬取n页
n = config.getn()

# 打开excel文件
wb = openpyxl.load_workbook('data.xlsx')
try:
    # 选定表格
    ws = wb["招标公告"]
    # 删除sheet
    wb.remove(ws)
    ws = wb["招标预告"]
    wb.remove(ws)
    ws = wb["招标变更"]
    wb.remove(ws)
    ws = wb["中标结果"]
    wb.remove(ws)
except Exception as e:
    print(e)

'''
招标公告
'''
# 新建一个sheet
ws=wb.create_sheet('招标公告')
# 设置列宽
ws.column_dimensions['B'].width = 65.75
ws.column_dimensions['C'].width = 23
ws.column_dimensions['D'].width = 52.38
ws.append(["省份", "招标信息", "日期", "招标网址"])

# 爬取n页
for k in range(n):
    url = "https://www.bidcenter.com.cn/zbpage-1-" + str(k) + ".html"
    f = requests.get(url)
    soup = BeautifulSoup(f.content, "lxml")
    print("正在爬取第" + str(k+1) + "页")
    for k in soup.find_all(id="searchcontent"):
        # print(k)
        for li in k.find_all('li'):
            for left in li.find_all('div', class_="s_c_l_left"):
                # 招标信息

                a = left.find_all('a')
                # 招标日期
                right = li.find_all('div', class_="s_c_l_right")

                # 省份，信息，日期, 网址
                line = [config.fix(a[0].string), config.fix(a[1].string), config.fix(right[0].get_text()), "https://www.bidcenter.com.cn" + a[1]['href']]
                # 写入
                ws.append(line)


'''
招标预告
'''
# 新建一个sheet
ws=wb.create_sheet('招标预告')
# 设置列宽
ws.column_dimensions['B'].width = 65.75
ws.column_dimensions['C'].width = 23
ws.column_dimensions['D'].width = 52.38
ws.append(["省份", "招标信息", "日期", "招标网址"])


# 爬取n页
for k in range(n):
    url = "https://www.bidcenter.com.cn/zbpage-2-" + str(k) + ".html"
    f = requests.get(url)
    soup = BeautifulSoup(f.content, "lxml")
    print("正在爬取第" + str(k+1) + "页")
    for k in soup.find_all(id="searchcontent"):
        # print(k)
        for li in k.find_all('li'):
            for left in li.find_all('div', class_="s_c_l_left"):
                # 招标信息

                a = left.find_all('a')
                # 招标日期
                right = li.find_all('div', class_="s_c_l_right")

                # 省份，信息，日期, 网址
                line = [config.fix(a[0].string), config.fix(a[1].string), config.fix(right[0].get_text()), "https://www.bidcenter.com.cn" + a[1]['href']]
                # 写入
                ws.append(line)


'''
招标变更
'''
# 新建一个sheet
ws=wb.create_sheet('招标变更')
# 设置列宽
ws.column_dimensions['B'].width = 65.75
ws.column_dimensions['C'].width = 23
ws.column_dimensions['D'].width = 52.38
ws.append(["省份", "招标信息", "日期", "招标网址"])


# 爬取n页
for k in range(n):
    url = "https://www.bidcenter.com.cn/zbpage-6-" + str(k) + ".html"
    f = requests.get(url)
    soup = BeautifulSoup(f.content, "lxml")
    print("正在爬取第" + str(k+1) + "页")
    for k in soup.find_all(id="searchcontent"):
        # print(k)
        for li in k.find_all('li'):
            for left in li.find_all('div', class_="s_c_l_left"):
                # 招标信息

                a = left.find_all('a')
                # 招标日期
                right = li.find_all('div', class_="s_c_l_right")

                # 省份，信息，日期, 网址
                line = [config.fix(a[0].string), config.fix(a[1].string), config.fix(right[0].get_text()), "https://www.bidcenter.com.cn" + a[1]['href']]
                # 写入
                ws.append(line)


'''
中标结果
'''
# 新建一个sheet
ws=wb.create_sheet('中标结果')
# 设置列宽
ws.column_dimensions['B'].width = 65.75
ws.column_dimensions['C'].width = 23
ws.column_dimensions['D'].width = 52.38
ws.append(["省份", "招标信息", "日期", "招标网址"])

# 爬取n页
for k in range(n):
    url = "https://www.bidcenter.com.cn/zbpage-4-" + str(k) + ".html"
    f = requests.get(url)
    soup = BeautifulSoup(f.content, "lxml")
    print("正在爬取第" + str(k+1) + "页")
    for k in soup.find_all(id="searchcontent"):
        # print(k)
        for li in k.find_all('li'):
            for left in li.find_all('div', class_="s_c_l_left"):
                # 招标信息

                a = left.find_all('a')
                # 招标日期
                right = li.find_all('div', class_="s_c_l_right")

                # 省份，信息，日期, 网址
                line = [config.fix(a[0].string), config.fix(a[1].string), config.fix(right[0].get_text()), "https://www.bidcenter.com.cn" + a[1]['href']]
                # 写入
                ws.append(line)




# 保存excel文件
wb.save("data.xlsx")

