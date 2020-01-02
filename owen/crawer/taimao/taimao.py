'''
泰茂医疗器械网
http://www.ylqxzb.com/zbPublish/Index_n1.html
'''
from bs4 import BeautifulSoup
from lxml import html
import xml
import requests
from openpyxl import Workbook
import openpyxl
import os, sys
sys.path.insert(0, '../')
import config
# 定义爬取n页
n = config.getn()
# 去除多余

# 打开excel文件
wb = openpyxl.load_workbook('data.xlsx')

try:
    # 选定表格
    ws1 = wb["泰茂"]
    # 删除sheet
    wb.remove(ws1)
    ws2 = wb["集中标讯动态（本月）"]
    wb.remove(ws2)
    ws3 = wb["开标中标公示（本季）"]
    wb.remove(ws3)
except Exception as e:
    print(e)

'''
泰茂
'''
# 新建一个sheet
ws1=wb.create_sheet('泰茂')
# 设置列宽
ws1.column_dimensions['B'].width = 110.88
ws1.column_dimensions['C'].width = 14.63
ws1.column_dimensions['D'].width = 62.25
ws1.append(["序号", "招标信息", "日期", "招标网址"])

# 爬取n页
for k in range(1, n):
    url = "http://www.ylqxzb.com/zbPublish/Index_n" + str(k) + ".html"
    f = requests.get(url)
    soup = BeautifulSoup(f.content, "lxml")
    print("正在爬取第" + str(k) + "页")
    for table in soup.find_all(class_="ta"):
    #     # print(k)
        for tr in table.find_all('tr'):
            td = tr.find_all('td')
            a = tr.find_all('a')
            if(len(td) == 0):
                continue
            # 序号，标题，发布时间，网址
            line = [config.fix(td[0].string), config.fix(a[0].string), config.fix(td[2].string), "http://www.ylqxzb.com/zbPublish/" + a[0]['href']]
            ws1.append(line)

'''
集中标讯动态（本月）
'''
# 新建一个sheet
ws2=wb.create_sheet('集中标讯动态（本月）')
# 设置列宽
ws2.column_dimensions['B'].width = 110.88
ws2.column_dimensions['C'].width = 14.63
ws2.column_dimensions['D'].width = 62.25
ws2.append(["序号", "招标信息", "日期", "招标网址"])

# 爬取n页
for k in range(1, n):
    url = "http://www.ylqxzb.com/zbPublish/Index_u3_n" + str(k) + ".html"
    f = requests.get(url)
    soup = BeautifulSoup(f.content, "lxml")
    print("正在爬取第" + str(k) + "页")
    for table in soup.find_all(class_="ta"):
        # print(table)
        for tr in table.find_all('tr'):
            td = tr.find_all('td')
            a = tr.find_all('a')
            if(len(td) == 0):
                continue
            # 序号，标题，发布时间，网址
            line = [config.fix(td[0].string), config.fix(a[0].string), config.fix(td[2].string), "http://www.ylqxzb.com/zbPublish/" + a[0]['href']]
            # 写入
            ws2.append(line)


'''
开标中标公示（本季）
'''
# 新建一个sheet
ws3=wb.create_sheet('开标中标公示（本季）')
# 设置列宽
ws3.column_dimensions['B'].width = 11.88
ws3.column_dimensions['C'].width = 71.38
ws3.column_dimensions['D'].width = 10.38
ws3.column_dimensions['E'].width = 14.5
ws3.column_dimensions['F'].width = 81.13
ws3.append(["序号", "类别", "项目", "项目状态", "上传时间", "网址"])

# 爬取n页
for k in range(1, n):
    url = "http://www.ylqxzb.com/zbProjectFit/Index_e4_n" + str(k) + ".html"
    f = requests.get(url)
    soup = BeautifulSoup(f.content, "lxml")
    print("正在爬取第" + str(k) + "页")
    for table in soup.find_all(class_="ta"):
        # print(table)
        for tr in table.find_all('tr'):
            td = tr.find_all('td')
            a = tr.find_all('a')
            if(len(td) == 0):
                continue
            # 序号，类别，项目，项目状态， 上传时间， 网址
            line = [config.fix(td[0].string), config.fix(td[1].string), config.fix(a[0].string), config.fix( td[3].find_all("span")[0].string ), config.fix(td[4].string), "http://www.ylqxzb.com/zbPublish/" + a[0]['href']]
            # 写入
            ws3.append(line)

# 保存excel文件
wb.save("data.xlsx")

